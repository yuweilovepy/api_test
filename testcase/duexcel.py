#- *- coding:utf-8 -*-

from openpyxl import load_workbook

"""
使用面向对象的
"""

# 使用面向对象的方法
class Case:
    def __init__(self):
        self.case_id=None
        self.title = None
        self.a=None
        self.b=None
        self.excepted=None

class DoExcel:
    def __init__(self,book_nmae):
        self.book_name=book_nmae


    def get_sheet(self,sheet_name):
        #读取表数据
        wb=load_workbook(self.book_name)
        sheet=wb[sheet_name]
        cases=[] #存储 对象
        for i in range(2,sheet.max_row+1):
            case = Case()  # 使用面向对象的方法
            case.case_id=sheet.cell(i,1).value
            case.title=sheet.cell(i,2).value
            case.a=sheet.cell(i,3).value
            case.b=sheet.cell(i,4).value
            case.excepted=sheet.cell(i,5).value
            cases.append(case)
        # for case in cases:
        #     print(case.case_id)
        return cases


    def set_sheet(self,row,col,value,sheet_name):
        #写入表数据
        wb = load_workbook(self.book_name)
        sheet = wb[sheet_name]
        sheet.cell(row,col,value)
        wb.save(self.book_name)  #  保存并关闭
        wb.close()

if __name__ == '__main__':
    DoExcel("case_data.xlsx").get_sheet("add")